from openpyxl.styles import PatternFill
import openpyxl
from investmentbot.utils import *
from investmentbot.define import *
import logging
import pandas as pd

class Excel:
    def __init__(self, file_name):
        self.file_name = file_name
    
    def _header(self):
        header = ['日期', 
                  '目標市值', 
                  '股票市值', 
                  '現金', 
                  '股票數量', 
                  '成本', 
                  'Current State', 
                  'TP1 Counter', 
                  'TP2 Counter',
                  '等入市-上月高於sma',
                  '現金投入',
                  'Target',
                  '操作',
                  '總資產',
                  '現金比率',
                  '回報',
                  'MDD',
                  '上月sma',
                  '上月Underlying Close',
                  '上日Underlying Close',
                  '當日Underlying Open',
                  '當日Deriv Open']
        return header

    def _color(self):
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        market_state_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        account_state_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        operation_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        stat_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        result = {}
        result['日期'] = header_fill
        result['目標市值'] = account_state_fill
        result['股票市值'] = account_state_fill
        result['現金'] = account_state_fill
        result['股票數量'] = account_state_fill
        result['成本']  = account_state_fill
        result['Current State'] = account_state_fill
        result['TP1 Counter'] = account_state_fill
        result['TP2 Counter'] = account_state_fill
        result['等入市-上月高於sma'] = account_state_fill
        result['現金投入'] = account_state_fill
        result['Target'] = operation_fill
        result['操作'] = operation_fill
        result['總資產'] = stat_fill
        result['現金比率'] = stat_fill
        result['回報'] = stat_fill
        result['MDD'] = stat_fill
        result['上月sma'] = market_state_fill
        result['上月Underlying Close'] = market_state_fill
        result['上日Underlying Close'] = market_state_fill
        result['當日Underlying Open'] = market_state_fill
        result['當日Deriv Open'] = market_state_fill

        return result

    def _format_float(self, number, decimal):
        return round(number, decimal)

    def _to_dict(self, event):
        total_value = event.stock + event.cash
        result = {}
        result['日期'] = int(event.date)
        result['目標市值'] = self._format_float(event.target_value, 2)
        result['股票市值'] = self._format_float(event.stock, 2)
        result['現金'] = self._format_float(event.cash, 2)
        result['股票數量'] = event.number_of_stocks
        result['成本']  = self._format_float(event.cost, 2)
        result['Current State'] = event.state.name
        result['TP1 Counter'] = event.tp1_counter
        result['TP2 Counter'] = event.tp2_counter
        result['等入市-上月高於sma'] = event.wait_in_and_previous_is_higher_than_sma
        result['現金投入'] = event.cash_in
        result['Target'] = event.action.name
        result['操作'] = event.operation
        result['總資產'] = self._format_float(total_value, 2)
        result['現金比率'] = f"{self._format_float(event.cash / total_value * 100, 2)}%"
        result['回報'] = f"{self._format_float((total_value - event.cost) / event.cost * 100, 2)}%"
        result['MDD'] = f"{self._format_float(event.mdd, 2)}%"
        result['上月sma'] = self._format_float(event.previous_month_sma, 2)
        result['上月Underlying Close'] = self._format_float(event.previous_month_close, 2)
        result['上日Underlying Close'] = self._format_float(event.previous_day_close, 2)
        result['當日Underlying Open'] = self._format_float(event.underlying_price_open, 2)
        result['當日Deriv Open'] = self._format_float(event.deriv_price_open, 2)
        return result

    def to_file(self, events, params, stat=None, with_header=False):
        header = self._header()
        color_dict = self._color()
        
        try:
            wb = openpyxl.load_workbook(self.file_name)
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active

        if with_header:
            ws.append(header)
            for cell in ws[1]:
                cell.fill = color_dict[cell.value]

        for event in events:
            event_dict = self._to_dict(event)
            row = [event_dict[key] for key in header]
            ws.append(row)

        ws.freeze_panes = 'A2'

        for column in ws.columns:  
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 10)  # Add some padding
            ws.column_dimensions[column_letter].width = adjusted_width

        sheet = None
        if 'Parameters' in wb.sheetnames:
            wb.remove(wb['Parameters'])
        
        sheet = wb.create_sheet(title='Parameters')
        for key, value in params.items():
            sheet.append([key, value])

        if stat != None:
            if 'Stat' in wb.sheetnames:
                wb.remove(wb['Stat'])
            
            sheet = wb.create_sheet(title='Stat')
            for key, value in stat.items():
                sheet.append([key, value])

        wb.save(self.file_name) 

    def from_file(self):
        try:
            data = pd.read_excel(self.file_name).to_dict(orient='records')[-1]
            return data

        except FileNotFoundError:
            logging.error(f"[E] File {self.file_name} does not exist. Exit")
            return None
    
    def clear_sheets(self):
        try:
            wb = openpyxl.load_workbook(self.file_name)
        except:
            return

        for sheet in wb.sheetnames:
            wb.remove(wb[sheet])

        sheet = wb.create_sheet(title='History')
        wb.active = sheet

        wb.save(self.file_name)