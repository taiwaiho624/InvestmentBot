from enum import Enum
from openpyxl.styles import PatternFill
import openpyxl

hkd_to_usd = 7.85

class State(Enum):
    WAIT_IN = 1
    WAIT_OUT = 2
    SL_ONE = 3

def get_state_from_string(state_string):
    try:
        return State[state_string.upper()] 
    except KeyError:
        raise ValueError(f"{state_string} is not a valid State")

class Action(Enum):
    BUY_UP_TO = 1
    ENRTRY = 2
    TAKE_PROF1= 3
    TAKE_PROF2 = 4
    SELL_HALF = 5
    SELL_ALL = 6
    NO_ACTION = 7

class Side(Enum):
    HOLD = 0
    BUY = 1
    SELL = 2

class Event:
    def __init__(self, action, operation_price, operation):
        #event
        self.date = None
        self.action = action
        self.operation_price = operation_price
        self.operation = operation
        self.cash_in = 0

        #account state
        self.target_value = None
        self.stock = None
        self.cash = None
        self.number_of_stocks = None
        self.cost = None
        self.state = None
        self.tp1_counter = None
        self.tp2_counter = None
        self.wait_in_and_previous_is_higher_than_sma = None

        #market state
        self.previous_month_sma = None
        self.previous_month_close = None
        self.underlying_price_open = None
        self.underlying_price_low = None
        self.underlying_price_high = None
        self.deriv_price_high = None
        self.deriv_price_open = None
        self.deriv_price_low = None
        
    def _format_float(self, number, decimal):
        return round(number, decimal)
    
    def header(self):
        header = ['日期', 
                  '目標市值', 
                  '股票市值', 
                  '現金', 
                  '股票數量', 
                  '成本', 
                  'Current State', 
                  'TP1 Counter', 
                  'TP2 Counter',
                  '等入市-上月高於sma',
                  '現金投入',
                  'Target',
                  '操作',
                  '總資產',
                  '現金比率',
                  '回報',
                  '上月sma',
                  '上月Underlying Close',
                  '當日Underlying Open',
                  '當日Underlying Low',
                  '當日Underlying High',
                  '當日Deriv High',
                  '當日Deriv Open',
                  '當日Deriv Low']
        return header
    
    def color(self):
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        market_state_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        account_state_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        operation_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        stat_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        result = {}
        result['日期'] = header_fill
        result['目標市值'] = account_state_fill
        result['股票市值'] = account_state_fill
        result['現金'] = account_state_fill
        result['股票數量'] = account_state_fill
        result['成本']  = account_state_fill
        result['Current State'] = account_state_fill
        result['TP1 Counter'] = account_state_fill
        result['TP2 Counter'] = account_state_fill
        result['等入市-上月高於sma'] = account_state_fill
        result['現金投入'] = account_state_fill
        result['Target'] = operation_fill
        result['操作'] = operation_fill
        result['總資產'] = stat_fill
        result['現金比率'] = stat_fill
        result['回報'] = stat_fill
        result['上月sma'] = market_state_fill
        result['上月Underlying Close'] = market_state_fill
        result['當日Underlying Open'] = market_state_fill
        result['當日Underlying Low'] = market_state_fill
        result['當日Underlying High'] = market_state_fill
        result['當日Deriv High'] = market_state_fill
        result['當日Deriv Open'] = market_state_fill
        result['當日Deriv Low'] = market_state_fill

        return result

    def to_dict(self):
        total_value = self.stock + self.cash
        result = {}
        result['日期'] = int(self.date)
        result['目標市值'] = self._format_float(self.target_value, 2)
        result['股票市值'] = self._format_float(self.stock, 2)
        result['現金'] = self._format_float(self.cash, 2)
        result['股票數量'] = self.number_of_stocks
        result['成本']  = self._format_float(self.cost, 2)
        result['Current State'] = self.state.name
        result['TP1 Counter'] = self.tp1_counter
        result['TP2 Counter'] = self.tp2_counter
        result['等入市-上月高於sma'] = self.wait_in_and_previous_is_higher_than_sma
        result['現金投入'] = self.cash_in
        result['Target'] = self.action.name
        result['操作'] = self.operation
        result['總資產'] = self._format_float(total_value, 2)
        result['現金比率'] = f"{self._format_float(self.cash / total_value * 100, 2)}%"
        result['回報'] = f"{self._format_float((total_value - self.cost) / self.cost * 100, 2)}%"
        result['上月sma'] = self._format_float(self.previous_month_sma, 2)
        result['上月Underlying Close'] = self._format_float(self.previous_month_close, 2)
        result['當日Underlying Open'] = self._format_float(self.underlying_price_open, 2)
        result['當日Underlying Low'] = self._format_float(self.underlying_price_low, 2)
        result['當日Underlying High'] = self._format_float(self.underlying_price_high, 2)
        result['當日Deriv High'] = self._format_float(self.deriv_price_high, 2)
        result['當日Deriv Open'] = self._format_float(self.deriv_price_open, 2)
        result['當日Deriv Low'] = self._format_float(self.deriv_price_low, 2)
        return result

    def to_file(self, file_name):
        header = self.header()
        color_dict = self.color()
        event_dict = self.to_dict()
        row = [event_dict[key] for key in header]

        try:
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active

            ws.append(header)
            
            for cell in ws[1]:
                cell.fill = color_dict[cell.value]

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 10)  # Add some padding
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.append(row)
        wb.save(file_name) 

    def __str__(self):
        event_dict = self.to_dict()
        result = ""
        for key, item in event_dict.items():
            result += f"{key}={item} "
        return result